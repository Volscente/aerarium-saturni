import io
import re
import zipfile
from datetime import date
from decimal import Decimal
from pathlib import Path
from typing import BinaryIO

import httpx
import openpyxl
from sqlalchemy import select, update
from sqlalchemy.ext.asyncio import AsyncSession

from backend.models import EtfHolding

XlsxSource = Path | BinaryIO

GERMAN_MONTHS = {
    "Januar": 1,
    "Februar": 2,
    "März": 3,
    "April": 4,
    "Mai": 5,
    "Juni": 6,
    "Juli": 7,
    "August": 8,
    "September": 9,
    "Oktober": 10,
    "November": 11,
    "Dezember": 12,
}

ISSUER_BY_TICKER = {
    "EUNL": "ishares",
    "VWCE": "vanguard",
    "LYP6": "amundi",
}

GERMAN_COUNTRY_TO_ISO = {
    "Australien": "AU",
    "Belgien": "BE",
    "Deutschland": "DE",
    "Dänemark": "DK",
    "Finnland": "FI",
    "Frankreich": "FR",
    "Hongkong": "HK",
    "Irland": "IE",
    "Israel": "IL",
    "Italien": "IT",
    "Japan": "JP",
    "Kanada": "CA",
    "Neuseeland": "NZ",
    "Niederlande": "NL",
    "Norwegen": "NO",
    "Portugal": "PT",
    "Schweden": "SE",
    "Schweiz": "CH",
    "Singapur": "SG",
    "Spanien": "ES",
    "Vereinigte Staaten": "US",
    "Vereinigtes Königreich": "GB",
    "Österreich": "AT",
}

OPENFIGI_MAPPING_URL = "https://api.openfigi.com/v3/mapping"
_OPENFIGI_JOBS_PER_REQUEST = 10

_ISIN_PATTERN = re.compile(r"[A-Z0-9]{12}")
_GERMAN_DATE_PATTERN = re.compile(r"(\d{1,2})\.\s*([A-Za-zÀ-ÿ]+)\s*(\d{4})")
_AMUNDI_DATE_PATTERN = re.compile(r"zum (\d{2})/(\d{2})/(\d{4})")
_LEADING_TICKER_PATTERN = re.compile(r"[A-Za-z0-9]+")


def convert_holdings_xlsx(source: XlsxSource, ticker: str | None = None) -> list[dict[str, str]]:
    """Convert an issuer holdings XLSX export into EtfHoldingRow-shaped dict rows.

    Dispatches on the issuer's ticker rather than sniffing the file's content,
    since issuer layouts are structurally different enough that a wrong guess
    would silently misparse columns rather than fail loudly.

    Args:
        source: Path to the issuer XLSX export, or an open binary file-like
            object (e.g. the bytes of an in-memory upload).
        ticker: The issuer's ticker, or a filename/stem carrying it as a
            leading token (e.g. ``"EUNL (1)"``, ``"EUNL_2026-07-23"``, both
            resolve to ``EUNL``) — repeat downloads of the same file get
            suffixed by the browser, so an exact match would be too brittle
            for routine re-uploads. Case-insensitive. Defaults to
            ``source.stem`` when ``source`` is a ``Path``; required otherwise,
            since a file-like object has no filename to infer it from.

    Returns:
        List of dicts with string values ready to be written as CSV rows or
        passed straight into ``EtfHoldingRow``. Each dict carries whichever
        identifier the issuer actually publishes — ``stock_ticker`` for
        iShares/Vanguard, ``stock_isin`` for Amundi — never both, so a blank
        cell for the other identifier is never produced (``EtfHoldingRow``
        only normalises a blank/absent *ISIN* to ``None``; a blank
        ``stock_ticker`` would fail its ``min_length=1`` constraint). Each
        dict also carries ``stock_country`` (an ISO 3166-1 alpha-2 code, or
        ``None`` if it couldn't be derived) — this field is not part of
        ``EtfHoldingRow`` and is read directly off the dict by the upload
        handler when constructing ``EtfHolding`` rows. Cash, money-market,
        derivative, and zero-weight lines are dropped, since ``EtfHoldingRow``
        models discrete stock constituents only and requires
        ``weight_percentage > 0``.

    Raises:
        ValueError: If ``ticker`` is omitted and ``source`` isn't a ``Path``,
            or the resolved ticker has no registered converter.
    """
    if ticker is None:
        if not isinstance(source, Path):
            raise ValueError("ticker must be provided when source is not a file path")
        ticker = source.stem
    match = _LEADING_TICKER_PATTERN.match(ticker)
    resolved_ticker = match.group(0).upper() if match else ticker.upper()
    issuer = ISSUER_BY_TICKER.get(resolved_ticker)
    if issuer is None:
        raise ValueError(f"Unrecognised ETF ticker '{resolved_ticker}'; no converter registered")
    return _CONVERTERS[issuer](source)


def write_csv(rows: list[dict[str, str]], output_path: Path) -> None:
    """Write converted holding rows to a CSV file matching EtfHoldingRow's field names.

    The header is taken from the first row's keys, since each issuer's rows
    carry only the identifier column it actually has data for (see
    ``convert_holdings_xlsx``); writing an unused identifier column with
    blank cells would fail ``EtfHoldingRow`` validation on upload.

    Args:
        rows: Rows produced by ``convert_holdings_xlsx``; must be non-empty.
        output_path: Destination CSV path; parent directory must already exist.
    """
    import csv

    with output_path.open("w", newline="", encoding="utf-8") as f:
        writer = csv.DictWriter(f, fieldnames=list(rows[0].keys()))
        writer.writeheader()
        writer.writerows(rows)


async def resolve_stock_isin_aliases(
    session: AsyncSession,
    http_client: httpx.AsyncClient,
) -> int:
    """Backfill stock_isin on ticker-only EtfHolding rows via OpenFIGI.

    OpenFIGI's mapping API only resolves ISIN -> ticker, never the reverse, so
    this queries every distinct stock_isin already stored in etf_holdings
    (i.e. every Amundi-sourced ISIN seen so far, from any upload), resolves
    each to its ticker via POST /v3/mapping (idType=ID_ISIN, batched up to the
    API's per-request job limit), then updates any row across ALL ETFs still
    missing stock_isin whose (stock_ticker, stock_country) matches the
    resolved (ticker, country) pair. Matching on country as well as ticker
    avoids merging unrelated companies that coincidentally share a ticker on
    different exchanges. Re-running this after every upload (regardless of
    issuer or order) is idempotent: the UPDATE's `stock_isin IS NULL` clause
    is a no-op for rows already backfilled in a previous run.

    Args:
        session: Async SQLAlchemy session.
        http_client: Shared httpx.AsyncClient used to call OpenFIGI.

    Returns:
        Number of EtfHolding rows whose stock_isin was backfilled by this call.

    Raises:
        httpx.HTTPError: If OpenFIGI is unreachable. Callers should treat this
            as non-fatal to the upload itself — the holdings are already
            committed; reconciliation can be retried on the next upload.
    """
    isins_result = await session.execute(
        select(EtfHolding.stock_isin).where(EtfHolding.stock_isin.is_not(None)).distinct()
    )
    isins = [row[0] for row in isins_result.all()]
    if not isins:
        return 0

    resolved: dict[str, str] = {}
    for i in range(0, len(isins), _OPENFIGI_JOBS_PER_REQUEST):
        batch = isins[i : i + _OPENFIGI_JOBS_PER_REQUEST]
        jobs = [{"idType": "ID_ISIN", "idValue": isin} for isin in batch]
        response = await http_client.post(OPENFIGI_MAPPING_URL, json=jobs)
        response.raise_for_status()
        for isin, mapping_result in zip(batch, response.json()):
            data = mapping_result.get("data")
            ticker = data[0].get("ticker") if data else None
            if ticker:
                resolved[isin] = ticker

    updated_rows = 0
    for isin, ticker in resolved.items():
        result = await session.execute(
            update(EtfHolding)
            .where(
                EtfHolding.stock_ticker == ticker,
                EtfHolding.stock_country == _country_from_isin(isin),
                EtfHolding.stock_isin.is_(None),
            )
            .values(stock_isin=isin)
        )
        updated_rows += result.rowcount or 0
    if updated_rows:
        await session.commit()
    return updated_rows


def _load_workbook(source: XlsxSource):
    """Load an XLSX workbook, tolerating issuer files with malformed style XML.

    Amundi exports have been observed with an invalid aRGB color value
    (``rgb="0xffffff"`` instead of a bare hex string) in ``xl/styles.xml``,
    which makes openpyxl refuse to open the file at all. When that happens,
    the stylesheet is rewritten in memory before retrying — styling is
    irrelevant here, only cell values matter.

    Args:
        source: Path to the XLSX file, or an open binary file-like object.

    Returns:
        The loaded ``openpyxl.Workbook``.
    """
    if not isinstance(source, Path):
        source.seek(0)
    try:
        return openpyxl.load_workbook(source, data_only=True)
    except ValueError:
        if not isinstance(source, Path):
            source.seek(0)
        return openpyxl.load_workbook(_sanitize_invalid_colors(source), data_only=True)


def _sanitize_invalid_colors(source: XlsxSource) -> io.BytesIO:
    """Rewrite non-hex ``rgb="0x..."`` color values in a workbook's styles.xml.

    Args:
        source: Path or open binary file-like object of the XLSX file whose
            stylesheet openpyxl rejected.

    Returns:
        An in-memory copy of the XLSX zip archive with valid aRGB hex values.
    """
    buf = io.BytesIO()
    with zipfile.ZipFile(source) as zin, zipfile.ZipFile(buf, "w") as zout:
        for item in zin.infolist():
            data = zin.read(item.filename)
            if item.filename == "xl/styles.xml":
                text = data.decode("utf-8")
                text = re.sub(
                    r'rgb="0x([0-9A-Fa-f]+)"',
                    lambda m: f'rgb="{m.group(1).upper().rjust(8, "F")}"',
                    text,
                )
                data = text.encode("utf-8")
            zout.writestr(item, data)
    buf.seek(0)
    return buf


def _parse_german_date(text: str) -> date:
    """Parse a German day/month-name/year date, with or without separating spaces.

    Args:
        text: Text containing a date such as ``"23.Juli2026"`` or
            ``"Per 30. Juni 2026"``; only the date portion is matched.

    Returns:
        The parsed date.

    Raises:
        ValueError: If no date pattern is found, or the month name is unknown.
    """
    match = _GERMAN_DATE_PATTERN.search(text)
    if not match:
        raise ValueError(f"Cannot parse German date from '{text}'")
    day, month_name, year = match.groups()
    if month_name not in GERMAN_MONTHS:
        raise ValueError(f"Unknown German month name '{month_name}'")
    return date(int(year), GERMAN_MONTHS[month_name], int(day))


def _clean_weight(value: object) -> Decimal | None:
    """Parse a weight value into a percentage Decimal, or None if not strictly positive.

    Args:
        value: Raw weight cell value — a float/int already in percentage
            units, or a German-formatted string like ``"4,4503\xa0%"``.

    Returns:
        The weight as a percentage, quantized to 4 decimal places, or ``None``
        if the value is missing, unparseable, or not strictly positive.
    """
    if value is None:
        return None
    if isinstance(value, str):
        cleaned = value.replace("\xa0", "").replace("%", "").replace(",", ".").strip()
        if not cleaned:
            return None
        try:
            weight = Decimal(cleaned)
        except Exception:
            return None
    else:
        weight = Decimal(str(value))
    quantized = weight.quantize(Decimal("0.0001"))
    if quantized <= 0:
        return None
    return quantized


def _country_from_isin(isin: str) -> str:
    """Extract the ISO 3166-1 alpha-2 country code from an ISIN's first two characters.

    Args:
        isin: A 12-character ISIN, e.g. "DE0007037129".

    Returns:
        The two-letter country code, e.g. "DE". No validation beyond slicing —
        callers already validate ISIN format upstream (see EtfHoldingRow / the
        Amundi converter's _ISIN_PATTERN check).
    """
    return isin[:2]


def _german_country_to_iso(country_name: str) -> str | None:
    """Map a German country name from an iShares export to its ISO alpha-2 code.

    Backed by a small static dict covering the country names observed in
    EUNL.xlsx's Standort column. Extend the dict as new countries appear in
    future holdings uploads.

    Args:
        country_name: The raw Standort cell value.

    Returns:
        The ISO alpha-2 code, or None if the country name isn't in the
        static mapping — the row still converts, just without stock_country.
    """
    return GERMAN_COUNTRY_TO_ISO.get(country_name)


def _convert_ishares(source: XlsxSource) -> list[dict[str, str]]:
    """Convert an iShares holdings export (e.g. EUNL.xlsx) to holding rows.

    Header row 3 (German column names); data from row 4. Only rows with
    ``Anlageklasse == "Aktien"`` are kept — cash, money-market, and
    derivative overlay lines carry small nonzero weights and would otherwise
    pass the weight filter despite not being stock constituents.

    Args:
        source: Path or open binary file-like object of the iShares XLSX export.

    Returns:
        Converted holding rows keyed by ``stock_ticker`` (no ISIN available),
        with ``stock_country`` derived from the ``Standort`` column via
        ``_german_country_to_iso`` (``None`` if the country name isn't mapped).
    """
    wb = _load_workbook(source)
    ws = wb.active
    snapshot_date = _parse_german_date(str(ws.cell(row=1, column=2).value))

    rows = []
    for (
        ticker,
        name,
        _sector,
        asset_class,
        _market_value,
        weight_pct,
        _nominal_value,
        _nominale,
        _kurs,
        location,
        *_rest,
    ) in (row[:11] for row in ws.iter_rows(min_row=4, values_only=True)):
        if asset_class != "Aktien" or not ticker or not name:
            continue
        weight = _clean_weight(weight_pct)
        if weight is None:
            continue
        rows.append(
            {
                "stock_ticker": str(ticker).strip(),
                "stock_name": str(name).strip(),
                "stock_country": _german_country_to_iso(str(location).strip()) if location else None,
                "weight_percentage": str(weight),
                "snapshot_date": snapshot_date.isoformat(),
            }
        )
    return rows


def _convert_vanguard(source: XlsxSource) -> list[dict[str, str]]:
    """Convert a Vanguard holdings export (e.g. VWCE.xlsx) to holding rows.

    Header row 7 (German column names); data from row 8. Rows with no ticker
    (secondary listings of an already-included company) are dropped, since
    Vanguard exports provide no ISIN fallback identifier.

    Args:
        source: Path or open binary file-like object of the Vanguard XLSX export.

    Returns:
        Converted holding rows keyed by ``stock_ticker`` (no ISIN available),
        with ``stock_country`` taken directly from the ``Region`` column
        (already ISO 3166-1 alpha-2, unlike iShares' German country names).
    """
    wb = _load_workbook(source)
    ws = wb.active
    title_row = " ".join(str(c) for c in next(ws.iter_rows(min_row=5, max_row=5, values_only=True)) if c)
    snapshot_date = _parse_german_date(title_row)

    rows = []
    for ticker, name, weight_pct, _sector, region, *_rest in (
        row[:5] for row in ws.iter_rows(min_row=8, values_only=True)
    ):
        if not ticker or not name:
            continue
        weight = _clean_weight(weight_pct)
        if weight is None:
            continue
        rows.append(
            {
                "stock_ticker": str(ticker).strip(),
                "stock_name": str(name).strip(),
                "stock_country": str(region).strip().upper() if region else None,
                "weight_percentage": str(weight),
                "snapshot_date": snapshot_date.isoformat(),
            }
        )
    return rows


def _convert_amundi(source: XlsxSource) -> list[dict[str, str]]:
    """Convert an Amundi holdings export (e.g. LYP6.xlsx) to holding rows.

    Metadata block precedes the real header row; the snapshot date is
    embedded in an "as of" label rather than a dedicated column. Weight is
    published as a fraction (e.g. ``0.0459`` for 4.59%) and is scaled ×100
    to match the other issuers' convention. Only rows with
    ``Anlageklasse == "EQUITY"`` are kept — futures overlay and cash lines
    carry ISIN-like identifiers too and would otherwise pass the identifier
    check despite not being stock constituents. Any row whose ISIN cell
    doesn't match the 12-alphanumeric pattern is skipped rather than treated
    as the end of the block — cash lines interleave with real holdings
    rather than trailing them, and the disclaimer footer follows all of it.

    Args:
        source: Path or open binary file-like object of the Amundi XLSX export.

    Returns:
        Converted holding rows keyed by ``stock_isin`` (no ticker available),
        with ``stock_country`` derived for free from the ISIN's own first two
        characters (``_country_from_isin``).
    """
    wb = _load_workbook(source)
    ws = wb.active
    rows_iter = ws.iter_rows(values_only=True)

    snapshot_date = None
    header_row_index = None
    for i, row in enumerate(rows_iter):
        if row[1] and "Verwaltetes Vermögen" in str(row[1]):
            match = _AMUNDI_DATE_PATTERN.search(str(row[1]))
            if match:
                day, month, year = match.groups()
                snapshot_date = date(int(year), int(month), int(day))
        if row[1] == "ISIN" and row[2] == "Name":
            header_row_index = i
            break
    if snapshot_date is None:
        raise ValueError("Cannot find snapshot date in Amundi workbook")
    if header_row_index is None:
        raise ValueError("Cannot find holdings header row in Amundi workbook")

    rows = []
    for row in rows_iter:
        isin, name, asset_class, weight_pct = row[1], row[2], row[3], row[5]
        if asset_class != "EQUITY" or not isin or not _ISIN_PATTERN.fullmatch(str(isin)) or not name:
            continue
        weight = _clean_weight(Decimal(str(weight_pct)) * 100 if weight_pct is not None else None)
        if weight is None:
            continue
        normalised_isin = str(isin).strip().upper()
        rows.append(
            {
                "stock_isin": normalised_isin,
                "stock_name": str(name).strip(),
                "stock_country": _country_from_isin(normalised_isin),
                "weight_percentage": str(weight),
                "snapshot_date": snapshot_date.isoformat(),
            }
        )
    return rows


_CONVERTERS = {
    "ishares": _convert_ishares,
    "vanguard": _convert_vanguard,
    "amundi": _convert_amundi,
}
